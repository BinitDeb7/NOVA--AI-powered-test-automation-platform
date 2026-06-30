 
from fpdf import FPDF
import datetime
import json
import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, GradientFill
import pdfkit
from markdown import markdown


# ─────────────────────────────────────────────
#  COLOUR PALETTE  — Premium Indigo & Midnight Theme
# ─────────────────────────────────────────────
C_NAVY         = (15,  30,  64)   # deep brand navy (kept for legacy Excel compatibility)
C_HEADER_BG    = (17,  24,  39)   # Slate-900 / Obsidian - primary header & cover block
C_PRIMARY      = (79,  70,  229)  # Iris / Indigo-600 - primary brand accent
C_PRIMARY_LIGHT= (165, 180, 252)  # Indigo-300 - lighter highlight
C_PRIMARY_SOFT = (245, 243, 255)  # Violet-50 - callout box backgrounds (AI advice)
C_SUCCESS      = (16,  185, 129)  # Emerald-500 - green success
C_SUCCESS_SOFT = (240, 253, 244)  # Emerald-50 - success step card background
C_FAIL         = (239,  68,  68)  # Red-500 - failure
C_FAIL_SOFT    = (254, 242, 242)  # Red-50 - failure step card background
C_WARN         = (245, 158,  11)  # Amber-500 - visual mismatch / warning
C_WARN_SOFT    = (255, 251, 235)  # Amber-50 - warning step card background
C_LIGHT_BG     = (249, 250, 251)  # Slate-50 - standard table header or panel bg
C_BORDER       = (229, 231, 235)  # Slate-200 - clean thin borders
C_TEXT_DARK    = (17,  24,  39)   # Slate-900 - dark readable text
C_TEXT_MID     = (107, 114, 128)  # Slate-500 - muted labels
C_TEXT_LIGHT   = (255, 255, 255)  # White text


def sanitize_text(text):
    """Replace non-latin-1 chars so fpdf doesn't crash."""
    if not text:
        return ""
    return str(text).encode('latin-1', 'replace').decode('latin-1')


def calculate_text_height(pdf, text, width, line_height):
    """Calculate the height that a multi_cell will take with a given width and line height."""
    if not text:
        return 0
    lines = 0
    # Split paragraphs by newline to preserve explicit formatting
    for paragraph in str(text).split('\n'):
        words = paragraph.split(' ')
        current_line = ""
        for word in words:
            if not word:
                continue
            test_line = current_line + " " + word if current_line else word
            if pdf.get_string_width(test_line) > width:
                lines += 1
                current_line = word
            else:
                current_line = test_line
        if current_line:
            lines += 1
    return lines * line_height


# ─────────────────────────────────────────────
#  BASE PDF CLASS  (shared header / footer)
# ─────────────────────────────────────────────

class TestVersePDF(FPDF):
    """Branded base class — Premium Indigo & Obsidian Theme."""

    def __init__(self, report_title="Test Report", *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.report_title = report_title
        self.set_auto_page_break(auto=True, margin=24)

    # ── HEADER ──────────────────────────────────────────────────────────
    def header(self):
        # Obsidian background bar
        self.set_fill_color(*C_HEADER_BG)
        self.rect(0, 0, 210, 16, 'F')
        # Iris left accent stripe
        self.set_fill_color(*C_PRIMARY)
        self.rect(0, 0, 4, 16, 'F')
        # Brand name
        self.set_xy(8, 3)
        self.set_font('Helvetica', 'B', 10)
        self.set_text_color(*C_TEXT_LIGHT)
        self.cell(60, 8, 'TestVerse  Nova', 0, 0, 'L')
        # Dot separator
        self.set_text_color(*C_PRIMARY)
        self.cell(4, 8, '|', 0, 0, 'C')
        # Report title (right)
        self.set_font('Helvetica', '', 8)
        self.set_text_color(156, 163, 175)  # grey-400
        self.cell(130, 8, sanitize_text(self.report_title), 0, 0, 'R')
        self.ln(19)

    # ── FOOTER ──────────────────────────────────────────────────────────
    def footer(self):
        self.set_y(-13)
        # Full-width subtle footer bar
        self.set_fill_color(249, 250, 251)  # slate-50
        self.rect(0, self.get_y() - 1, 210, 14, 'F')
        # Top border line
        self.set_draw_color(*C_BORDER)
        self.set_line_width(0.3)
        self.line(0, self.get_y() - 1, 210, self.get_y() - 1)
        # Footer text
        self.set_font('Helvetica', '', 7)
        self.set_text_color(*C_TEXT_MID)
        ts = datetime.datetime.now().strftime('%d %b %Y  %H:%M')
        self.set_x(12)
        self.cell(90, 8, f'TestVerse - Automated Test Report  |  {ts}', 0, 0, 'L')
        self.set_x(12)
        self.cell(186, 8, f'Page {self.page_no()}', 0, 0, 'R')

    # ── HELPERS ─────────────────────────────────────────────────────────
    def section_heading(self, title, icon=''):
        """Clean section heading with left primary accent bar."""
        self.ln(5)
        y = self.get_y()
        # Left accent bar in Iris
        self.set_fill_color(*C_PRIMARY)
        self.rect(12, y, 3, 8, 'F')
        # Heading text
        self.set_xy(18, y)
        self.set_font('Helvetica', 'B', 11)
        self.set_text_color(*C_HEADER_BG)
        self.cell(180, 8, sanitize_text(f'{icon}  {title}' if icon else title), 0, 1, 'L')
        # Subtle full-width rule
        self.set_draw_color(*C_BORDER)
        self.set_line_width(0.3)
        self.line(12, self.get_y(), 198, self.get_y())
        self.ln(4)
        self.set_text_color(*C_TEXT_DARK)

    def info_row(self, label, value, label_w=40, row_h=6.5):
        """Properly aligned label: value row."""
        self.set_x(14)
        self.set_font('Helvetica', 'B', 8)
        self.set_text_color(*C_TEXT_MID)
        self.cell(label_w, row_h, sanitize_text(label) + ':', 0, 0, 'L')
        self.set_font('Helvetica', '', 8)
        self.set_text_color(*C_TEXT_DARK)
        self.multi_cell(190 - label_w, row_h, sanitize_text(str(value)), 0, 'L')

    def status_chip(self, status_text, x, y, w=32, h=5.5):
        """Rounded status chip."""
        s = status_text.lower()
        if 'success' in s or 'pass' in s or 'match' in s or 'baseline' in s:
            bg = C_SUCCESS
        elif 'fail' in s or 'error' in s or 'mismatch' in s:
            bg = C_FAIL
        elif 'warn' in s or 'visual' in s:
            bg = C_WARN
        else:
            bg = C_PRIMARY
        self.set_fill_color(*bg)
        self.set_xy(x, y)
        self.set_font('Helvetica', 'B', 6.5)
        self.set_text_color(*C_TEXT_LIGHT)
        self.cell(w, h, sanitize_text(status_text.upper()), 0, 0, 'C', fill=True)


# ─────────────────────────────────────────────
#  TEST RUN PDF  (history & dashboard download)
# ─────────────────────────────────────────────

def _status_colors(status_text):
    """Return (solid_color, badge_bg, card_bg) for a step status."""
    s = status_text.lower()
    if 'success' in s or 'match' in s or 'baseline' in s:
        return C_SUCCESS, C_SUCCESS, C_SUCCESS_SOFT
    elif 'fail' in s or 'error' in s or 'mismatch' in s:
        return C_FAIL, C_FAIL, C_FAIL_SOFT
    elif 'warn' in s or 'visual' in s:
        return C_WARN, C_WARN, C_WARN_SOFT
    return C_PRIMARY, C_PRIMARY, C_PRIMARY_SOFT


def generate_pdf_report(url, results, filepath, summary=None):
    pdf = TestVersePDF(report_title="Automated Test Execution Report")
    pdf.add_page()

    # ═══════════════════════════════════════════════════════════════
    #  COVER SECTION
    # ═══════════════════════════════════════════════════════════════
    # Solid Obsidian hero block
    pdf.set_fill_color(*C_HEADER_BG)
    pdf.rect(0, 16, 210, 60, 'F')
    # Iris bottom stripe
    pdf.set_fill_color(*C_PRIMARY)
    pdf.rect(0, 76, 210, 2.5, 'F')
    # Decorative Iris geometric flair (extremely premium look)
    pdf.set_fill_color(129, 140, 248)  # Indigo-400
    pdf.ellipse(170, 20, 50, 50, 'F')
    pdf.set_fill_color(*C_PRIMARY)
    pdf.ellipse(178, 30, 34, 34, 'F')
    # Slate overlay to dim circles and keep text readable
    pdf.set_fill_color(*C_HEADER_BG)
    pdf.rect(0, 16, 155, 60, 'F')

    # Title
    pdf.set_xy(14, 22)
    pdf.set_font('Helvetica', 'B', 24)
    pdf.set_text_color(*C_TEXT_LIGHT)
    pdf.cell(140, 13, 'Test Execution Report', 0, 1, 'L')

    # Subtitle tag line
    pdf.set_xy(14, 37)
    pdf.set_font('Helvetica', '', 9.5)
    pdf.set_text_color(*C_PRIMARY_LIGHT)
    pdf.cell(140, 6, 'Nova QA Engine - Intelligent Web Automation & Visual Analysis', 0, 1, 'L')

    # Horizontal divider line inside cover
    pdf.set_draw_color(55, 65, 81)  # Slate-700
    pdf.set_line_width(0.3)
    pdf.line(14, 46, 110, 46)

    # Meta info rows inside cover
    ts_full = datetime.datetime.now().strftime('%A, %d %B %Y  at  %H:%M:%S')
    pdf.set_xy(14, 49)
    pdf.set_font('Helvetica', 'B', 7.5)
    pdf.set_text_color(156, 163, 175)  # grey-400
    pdf.cell(20, 5.5, 'Executed On:', 0, 0, 'L')
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(*C_TEXT_LIGHT)
    pdf.cell(120, 5.5, sanitize_text(ts_full), 0, 1, 'L')

    pdf.set_x(14)
    pdf.set_font('Helvetica', 'B', 7.5)
    pdf.set_text_color(156, 163, 175)
    pdf.cell(20, 5.5, 'Tested URL:', 0, 0, 'L')
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(*C_TEXT_LIGHT)
    pdf.multi_cell(130, 5.5, sanitize_text(url), 0, 'L')

    pdf.ln(10)   # clear the cover band

    # ═══════════════════════════════════════════════════════════════
    #  STATISTICS GRID
    # ═══════════════════════════════════════════════════════════════
    total  = len(results)
    passed = sum(1 for r in results
                 if str(r.get('status', '')).lower() in
                 ('success', 'new baseline', 'visuals match', 'successes'))
    failed = total - passed
    rate   = int((passed / total * 100) if total else 0)

    pdf.section_heading('Summary')

    # Optional AI summary block
    if summary:
        pdf.set_x(14)
        pdf.set_font('Helvetica', '', 7.5)
        text_h = calculate_text_height(pdf, sanitize_text(summary), 176, 5)
        box_h = 10.5 + text_h  # Title cell (5) + padding + margins
        
        sy = pdf.get_y()
        # Left accent block
        pdf.set_fill_color(*C_PRIMARY)
        pdf.rect(14, sy, 2, box_h, 'F')
        
        # Soft primary background box
        pdf.set_fill_color(*C_PRIMARY_SOFT)
        pdf.rect(16, sy, 182, box_h, 'F')
        
        # Render Title
        pdf.set_xy(20, sy + 2.5)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(*C_PRIMARY)
        pdf.cell(178, 5, 'AI Summary', 0, 1, 'L')
        
        # Render Text
        pdf.set_x(20)
        pdf.set_font('Helvetica', '', 7.5)
        pdf.set_text_color(*C_TEXT_DARK)
        
        for line in summary.split('\n'):
            line_str = sanitize_text(line).strip()
            if not line_str:
                pdf.ln(1)
                pdf.set_x(20)
                continue
            if line_str.startswith('#'):
                num_hashes = len(line_str) - len(line_str.lstrip('#'))
                clean_title = line_str.lstrip('#').strip()
                pdf.set_font('Helvetica', 'B', 8)
                pdf.multi_cell(176, 5, clean_title, 0, 'L')
                pdf.set_font('Helvetica', '', 7.5)
            else:
                pdf.multi_cell(176, 5, line_str, 0, 'L')
            pdf.set_x(20)
        
        # Reset Y to bottom of the summary box + margin
        pdf.set_y(sy + box_h + 4)

    # Stat cards row
    stat_y = pdf.get_y()
    card_w = 43
    card_h = 24
    gap = 3
    cards = [
        ('TOTAL STEPS', str(total),   C_HEADER_BG, C_TEXT_LIGHT),
        ('PASSED',      str(passed),  C_SUCCESS_SOFT, C_SUCCESS),
        ('FAILED',      str(failed),  C_FAIL_SOFT, C_FAIL),
        ('PASS RATE',   f'{rate}%',   C_PRIMARY_SOFT, C_PRIMARY),
    ]
    for i, (label, val, bg, fg) in enumerate(cards):
        cx = 14 + i * (card_w + gap)
        # Card background fill
        pdf.set_fill_color(*bg)
        pdf.rect(cx, stat_y, card_w, card_h, 'F')
        
        # Border
        pdf.set_draw_color(*C_BORDER)
        pdf.set_line_width(0.2)
        pdf.rect(cx, stat_y, card_w, card_h, 'D')
        
        # Value
        pdf.set_xy(cx, stat_y + 3)
        pdf.set_font('Helvetica', 'B', 18)
        pdf.set_text_color(*fg)
        pdf.cell(card_w, 11, val, 0, 0, 'C')
        
        # Label
        pdf.set_xy(cx, stat_y + 15)
        pdf.set_font('Helvetica', '', 6.5)
        if bg == C_HEADER_BG:
            pdf.set_text_color(156, 163, 175) # gray-400
        else:
            pdf.set_text_color(*C_TEXT_MID)
        pdf.cell(card_w, 6, label, 0, 0, 'C')

    pdf.set_y(stat_y + card_h + 4)

    # Pass-rate progress bar
    bar_x, bar_y_pos = 14, pdf.get_y()
    bar_total_w = 182
    bar_h = 5
    # Track (background)
    pdf.set_fill_color(*C_BORDER)
    pdf.rect(bar_x, bar_y_pos, bar_total_w, bar_h, 'F')
    # Fill
    filled_w = int(bar_total_w * rate / 100)
    if filled_w > 0:
        if rate >= 80:
            pdf.set_fill_color(*C_SUCCESS)
        elif rate >= 50:
            pdf.set_fill_color(*C_WARN)
        else:
            pdf.set_fill_color(*C_FAIL)
        pdf.rect(bar_x, bar_y_pos, filled_w, bar_h, 'F')
    # Label inside bar area
    pdf.set_xy(bar_x, bar_y_pos)
    pdf.set_font('Helvetica', 'B', 6)
    pdf.set_text_color(*C_TEXT_DARK)
    pdf.cell(bar_total_w, bar_h, f'  Pass Rate: {rate}%', 0, 1, 'L')
    pdf.ln(5)

    # Overall result banner (flat alert style)
    if failed == 0:
        pdf.set_fill_color(*C_SUCCESS_SOFT)
        pdf.set_text_color(*C_SUCCESS)
        pdf.set_draw_color(167, 243, 208) # green-200
        banner_text = '  ALL AUTOMATED TESTING VERIFICATIONS COMPLETED SUCCESSFULLY'
    else:
        pdf.set_fill_color(*C_FAIL_SOFT)
        pdf.set_text_color(*C_FAIL)
        pdf.set_draw_color(254, 205, 205) # red-200
        banner_text = f'  QA ENGINE ALERT - {failed} VERIFICATION STEP(S) DETECTED CRITICAL MISMATCH'
        
    pdf.set_x(14)
    pdf.set_font('Helvetica', 'B', 7.5)
    pdf.set_line_width(0.3)
    pdf.cell(182, 9, sanitize_text(banner_text), 1, 1, 'L', fill=True)
    pdf.ln(6)

    # ═══════════════════════════════════════════════════════════════
    #  STEP-BY-STEP CARDS
    # ═══════════════════════════════════════════════════════════════
    pdf.section_heading('Detailed Step Results', icon='')

    LABEL_W  = 28    # width of the label column (Command / Details / AI)
    VALUE_W  = 150   # remaining content width (182 - 32)
    CARD_X   = 14
    CARD_W   = 182
    INDENT_X = CARD_X + LABEL_W

    for idx, res in enumerate(results):
        step_num    = res.get('step') or (idx + 1)
        step_status = res.get('status', 'N/A')
        border_col, badge_bg, card_bg = _status_colors(step_status)

        command_text   = sanitize_text(res.get('command', '-'))
        details_text   = sanitize_text(res.get('details', ''))
        ai_text        = sanitize_text(res.get('ai_suggestion', ''))
        has_screenshot = bool(res.get('screenshot'))

        # ── Pre-calculate Card Heights ──
        card_header_h = 9
        
        # 1. Command block height
        cmd_h = calculate_text_height(pdf, command_text, VALUE_W, 5.5) + 3
        
        # 2. Details block height
        det_h = (calculate_text_height(pdf, details_text, VALUE_W, 5.5) + 2) if details_text else 0
        
        # 3. AI Suggestion block height
        ai_h = (calculate_text_height(pdf, ai_text, VALUE_W - 6, 5.5) + 6) if ai_text else 0
        
        # 4. Screenshot height
        ss_h = 82 if has_screenshot else 0
        
        card_body_h = cmd_h + det_h + ai_h + ss_h + 3  # total body height including card padding
        card_total_h = card_header_h + card_body_h

        # Page-break guard: if the card will overflow the page bottom, push it to a new page
        if pdf.get_y() + card_total_h > 270:
            pdf.add_page()

        card_top = pdf.get_y()

        # ── Draw Card Elements (Backgrounds & Borders) FIRST ──
        # 1. Card Header background
        pdf.set_fill_color(*card_bg)
        pdf.rect(CARD_X, card_top, CARD_W, card_header_h, 'F')

        # 2. Card Body background (white)
        pdf.set_fill_color(255, 255, 255)
        pdf.rect(CARD_X, card_top + card_header_h, CARD_W, card_body_h, 'F')

        # 3. Left indicator accent stripe
        pdf.set_fill_color(*border_col)
        pdf.rect(CARD_X, card_top, 3.5, card_total_h, 'F')

        # 4. Card outline frame border
        pdf.set_draw_color(*C_BORDER)
        pdf.set_line_width(0.25)
        pdf.rect(CARD_X, card_top, CARD_W, card_total_h, 'D')

        # ── Render Step Content ON TOP ──
        
        # Step Number (Left Header)
        pdf.set_xy(CARD_X + 5, card_top + 1.5)
        pdf.set_font('Helvetica', 'B', 8.5)
        pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(50, 6, f'Step {step_num}', 0, 0, 'L')

        # Status badge (Right Header)
        chip_w = 34
        chip_x = CARD_X + CARD_W - chip_w - 2.5
        chip_y = card_top + 1.7
        pdf.set_fill_color(*badge_bg)
        pdf.rect(chip_x, chip_y, chip_w, 5.5, 'F')
        pdf.set_xy(chip_x, chip_y)
        pdf.set_font('Helvetica', 'B', 6)
        pdf.set_text_color(*C_TEXT_LIGHT)
        pdf.cell(chip_w, 5.5, step_status.upper(), 0, 0, 'C')

        # Render rows inside Card Body
        current_y = card_top + card_header_h + 2

        # 1. Command Row
        pdf.set_xy(CARD_X + 5, current_y + 1.5)
        pdf.set_font('Helvetica', 'B', 7.5)
        pdf.set_text_color(*C_TEXT_MID)
        pdf.cell(LABEL_W - 5, 5.5, 'Command', 0, 0, 'L')
        
        pdf.set_xy(INDENT_X, current_y + 1.5)
        pdf.set_font('Helvetica', '', 7.5)
        pdf.set_text_color(*C_TEXT_DARK)
        pdf.multi_cell(VALUE_W, 5.5, command_text, 0, 'L')
        current_y = pdf.get_y() + 1.5

        # 2. Details Row (if exists)
        if details_text:
            pdf.set_xy(CARD_X + 5, current_y + 1)
            pdf.set_font('Helvetica', 'B', 7.5)
            pdf.set_text_color(*C_TEXT_MID)
            pdf.cell(LABEL_W - 5, 5.5, 'Details', 0, 0, 'L')
            
            pdf.set_xy(INDENT_X, current_y + 1)
            pdf.set_font('Helvetica', '', 7.5)
            pdf.set_text_color(*C_TEXT_DARK)
            pdf.multi_cell(VALUE_W, 5.5, details_text, 0, 'L')
            current_y = pdf.get_y() + 1.5

        # 3. AI Suggestion Box (if exists)
        if ai_text:
            ai_box_y = current_y + 1
            ai_box_h = calculate_text_height(pdf, ai_text, VALUE_W - 6, 5.5) + 4
            
            # Soft violet panel background
            pdf.set_fill_color(*C_PRIMARY_SOFT)
            pdf.rect(INDENT_X, ai_box_y, VALUE_W, ai_box_h, 'F')
            
            # Left solid Iris accent border
            pdf.set_fill_color(*C_PRIMARY)
            pdf.rect(INDENT_X, ai_box_y, 1.5, ai_box_h, 'F')
            
            # Text inside violet box
            pdf.set_xy(CARD_X + 5, ai_box_y + 2)
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_text_color(*C_PRIMARY)
            pdf.cell(LABEL_W - 5, 5.5, 'AI Advice', 0, 0, 'L')
            
            pdf.set_xy(INDENT_X + 4, ai_box_y + 2)
            pdf.set_font('Helvetica', 'I', 7)
            pdf.set_text_color(*C_TEXT_DARK)
            
            for line in ai_text.split('\n'):
                line_str = sanitize_text(line).strip()
                if not line_str:
                    pdf.ln(1)
                    pdf.set_x(INDENT_X + 4)
                    continue
                if line_str.startswith('#'):
                    num_hashes = len(line_str) - len(line_str.lstrip('#'))
                    clean_title = line_str.lstrip('#').strip()
                    pdf.set_font('Helvetica', 'BI', 7.5)
                    pdf.multi_cell(VALUE_W - 6, 5.5, clean_title, 0, 'L')
                    pdf.set_font('Helvetica', 'I', 7)
                else:
                    pdf.multi_cell(VALUE_W - 6, 5.5, line_str, 0, 'L')
                pdf.set_x(INDENT_X + 4)
                
            current_y = ai_box_y + ai_box_h + 1.5

        # 4. Screenshot block (if exists)
        if has_screenshot:
            try:
                img_data = base64.b64decode(res['screenshot'])
                img_file = io.BytesIO(img_data)
                
                img_label_y = current_y + 1.5
                pdf.set_xy(CARD_X + 5, img_label_y)
                pdf.set_font('Helvetica', 'B', 7.5)
                pdf.set_text_color(*C_TEXT_MID)
                pdf.cell(LABEL_W - 5, 5.5, 'Screenshot', 0, 1, 'L')
                
                img_h = 70
                img_w = 124
                img_y = img_label_y + 1.5
                img_x = CARD_X + (CARD_W - img_w) / 2  # center inside the card body
                
                # Render Image
                pdf.image(img_file, x=img_x, y=img_y, w=img_w, h=img_h)
                
                # Draw high-end subtle border frame around the image
                pdf.set_draw_color(*C_BORDER)
                pdf.set_line_width(0.2)
                pdf.rect(img_x, img_y, img_w, img_h, 'D')
                
                # Render subtitle caption below screenshot
                caption_y = img_y + img_h + 1.5
                pdf.set_xy(CARD_X, caption_y)
                pdf.set_font('Helvetica', 'I', 6.5)
                pdf.set_text_color(*C_TEXT_MID)
                pdf.cell(CARD_W, 5, f'Visual Regression Capture - Step {step_num}', 0, 1, 'C')
                
                current_y = caption_y + 4.5
            except Exception as err:
                pdf.set_xy(CARD_X + 5, current_y + 1.5)
                pdf.set_font('Helvetica', 'I', 7)
                pdf.set_text_color(*C_FAIL)
                pdf.cell(CARD_W, 5, f'[Screenshot render warning: {err}]', 0, 1, 'L')
                current_y = pdf.get_y() + 2

        # Move cursor y to the bottom of the card with small spacing
        pdf.set_y(card_top + card_total_h + 4.5)

    pdf.output(filepath)


# ─────────────────────────────────────────────
#  INLINE / DASHBOARD PDF  (export/pdf route)
# ─────────────────────────────────────────────

def generate_inline_pdf(summary, results, filepath, url=None):
    """Used by the /export/pdf dashboard route. Unifies layout with history downloads."""
    # Attempt to extract target URL from the first navigate command in results if not explicitly provided
    if not url or url == "N/A":
        url = "N/A"
        for r in results:
            cmd = str(r.get('command', '')).lower()
            if 'navigate' in cmd:
                parts = r.get('command', '').split('"')
                url = parts[1] if len(parts) > 1 else r.get('command', '').split()[-1]
                break
    generate_pdf_report(url, results, filepath, summary=summary)


# ─────────────────────────────────────────────
#  EXCEL REPORT
# ─────────────────────────────────────────────

def generate_excel_report(results, url, user_name, test_run_id):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Execution Report"

    # ── Styles ──────────────────────────────────────────────
    navy_fill    = PatternFill("solid", fgColor="0F1E40")
    accent_fill  = PatternFill("solid", fgColor="6378FF")
    success_fill = PatternFill("solid", fgColor="D1FAE5")
    fail_fill    = PatternFill("solid", fgColor="FEE2E2")
    warn_fill    = PatternFill("solid", fgColor="FEF3C7")
    alt_fill     = PatternFill("solid", fgColor="F0F3FF")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    header_row_fill = PatternFill("solid", fgColor="E8ECFF")

    white_bold   = Font(bold=True,  color="FFFFFF", size=11)
    navy_bold    = Font(bold=True,  color="0F1E40", size=11)
    accent_font  = Font(bold=True,  color="6378FF", size=10)
    muted_font   = Font(color="4B5563", size=9)
    body_font    = Font(color="111827", size=9)

    thin = Side(style='thin', color="D1D5DB")
    med  = Side(style='medium', color="6378FF")
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    accent_top   = Border(left=thin, right=thin, top=med, bottom=thin)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Title row ────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Nova — Test Execution Report'
    ws['A1'].font      = Font(bold=True, color="FFFFFF", size=16)
    ws['A1'].fill      = navy_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 36

    # ── Meta info ───────────────────────────────────────────
    meta = [
        ("Project",        "Nova Automated Testing"),
        ("Test Run ID",    str(test_run_id)),
        ("Executed By",    user_name),
        ("URL Tested",     url),
        ("Date / Time",    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws.merge_cells(f'A{i}:B{i}')
        ws[f'A{i}'] = k
        ws[f'A{i}'].font      = accent_font
        ws[f'A{i}'].alignment = left
        ws.merge_cells(f'C{i}:F{i}')
        ws[f'C{i}'] = v
        ws[f'C{i}'].font      = body_font
        ws[f'C{i}'].alignment = left
        ws.row_dimensions[i].height = 18

    # blank separator
    sep = len(meta) + 2
    ws.row_dimensions[sep].height = 8

    # ── Column headers ───────────────────────────────────────
    header_row = sep + 1
    headers = ["Step", "Command", "Status", "Details", "AI Suggestion", "Duration"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font      = Font(bold=True, color="0F1E40", size=9)
        c.fill      = header_row_fill
        c.alignment = center
        c.border    = accent_top
    ws.row_dimensions[header_row].height = 22

    # ── Data rows ────────────────────────────────────────────
    for r_idx, res in enumerate(results, start=header_row + 1):
        status  = str(res.get('status', ''))
        s_lower = status.lower()

        if 'success' in s_lower or 'match' in s_lower or 'baseline' in s_lower:
            s_fill = success_fill
            s_font = Font(bold=True, color="166534", size=9)
        elif 'fail' in s_lower or 'error' in s_lower:
            s_fill = fail_fill
            s_font = Font(bold=True, color="991B1B", size=9)
        else:
            s_fill = warn_fill
            s_font = Font(bold=True, color="92400E", size=9)

        row_fill = alt_fill if r_idx % 2 == 0 else white_fill

        data = [
            res.get('step', ''),
            res.get('command', ''),
            status,
            res.get('details', ''),
            res.get('ai_suggestion', ''),
            '',
        ]
        for col_idx, val in enumerate(data, 1):
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.border    = thin_border
            c.alignment = left if col_idx != 1 else center
            c.font      = body_font
            c.fill      = row_fill

        # Override status cell
        s_cell       = ws.cell(row=r_idx, column=3)
        s_cell.fill  = s_fill
        s_cell.font  = s_font
        s_cell.alignment = center

        ws.row_dimensions[r_idx].height = 20

    # ── Column widths ─────────────────────────────────────────
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 12

    # ── Summary row ──────────────────────────────────────────
    total  = len(results)
    passed = sum(1 for r in results if 'success' in str(r.get('status', '')).lower()
                 or 'match' in str(r.get('status', '')).lower()
                 or 'baseline' in str(r.get('status', '')).lower())
    failed = total - passed

    sum_row = header_row + total + 2
    ws.merge_cells(f'A{sum_row}:F{sum_row}')
    ws[f'A{sum_row}'] = f'Total: {total}   |   Passed: {passed}   |   Failed: {failed}   |   Pass Rate: {int(passed/total*100) if total else 0}%'
    ws[f'A{sum_row}'].font      = Font(bold=True, color="6378FF", size=10)
    ws[f'A{sum_row}'].alignment = center
    ws[f'A{sum_row}'].fill      = PatternFill("solid", fgColor="EEF1FF")
    ws.row_dimensions[sum_row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  CODE ANALYSIS PDF  (HTML → PDF via pdfkit)
# ─────────────────────────────────────────────

def generate_code_analysis_pdf(ai_markdown_output):
    """Generates a premium PDF from Markdown using pdfkit."""
    try:
        import pdfkit as _pdfkit_check  # availability check
    except ImportError:
        return None

    html_content = markdown(ai_markdown_output, extensions=['fenced_code', 'codehilite'])

    ts = datetime.datetime.now().strftime('%Y-%m-%d  %H:%M:%S')

    full_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AI Code Analysis Report – Nova</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&family=Fira+Code:wght@400;500&display=swap');

  * {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: 'Inter', 'Segoe UI', sans-serif;
    background: #ffffff;
    color: #111827;
    font-size: 10.5pt;
    line-height: 1.7;
    padding: 0;
  }}

  /* ── Cover band ── */
  .cover {{
    background: #0F1E40;
    color: #fff;
    padding: 36px 48px 28px;
    position: relative;
    border-bottom: 4px solid #6378FF;
  }}
  .cover .brand {{
    font-size: 11pt;
    color: #8a9fff;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 8px;
  }}
  .cover h1 {{
    font-size: 22pt;
    font-weight: 700;
    color: #fff;
    margin-bottom: 6px;
  }}
  .cover .sub {{
    font-size: 9pt;
    color: #a5b4fc;
  }}

  /* ── Body content ── */
  .content {{
    padding: 36px 48px;
  }}

  h1, h2, h3 {{
    color: #0F1E40;
    font-weight: 700;
    margin-top: 24px;
    margin-bottom: 8px;
    padding-bottom: 4px;
    border-bottom: 2px solid #6378FF;
  }}
  h1 {{ font-size: 16pt; }}
  h2 {{ font-size: 13pt; color: #6378FF; }}
  h3 {{ font-size: 11pt; color: #374151; border-bottom-color: #e5e7eb; }}

  p  {{ margin-bottom: 10px; }}
  ul, ol {{ margin-left: 20px; margin-bottom: 10px; }}
  li {{ margin-bottom: 4px; }}

  strong {{ color: #0F1E40; }}

  pre {{
    background: #1e293b;
    color: #e2e8f0;
    border-radius: 6px;
    padding: 14px 18px;
    font-family: 'Fira Code', 'Courier New', monospace;
    font-size: 9pt;
    overflow-x: auto;
    white-space: pre-wrap;
    word-break: break-word;
    margin: 12px 0;
    border-left: 4px solid #6378FF;
  }}

  code {{
    font-family: 'Fira Code', 'Courier New', monospace;
    font-size: 9pt;
    background: #f0f3ff;
    color: #3730a3;
    padding: 1px 5px;
    border-radius: 3px;
  }}
  pre code {{
    background: transparent;
    color: inherit;
    padding: 0;
  }}

  blockquote {{
    border-left: 4px solid #6378FF;
    background: #f0f3ff;
    padding: 10px 16px;
    margin: 12px 0;
    color: #374151;
    border-radius: 0 6px 6px 0;
  }}

  /* Codehilite (Atom One Dark) */
  .codehilite .c  {{ color: #5c6370; font-style: italic }}
  .codehilite .k  {{ color: #c678dd; font-weight: bold }}
  .codehilite .kn {{ color: #c678dd }}
  .codehilite .nf {{ color: #61afef }}
  .codehilite .nc {{ color: #e5c07b; font-weight: bold }}
  .codehilite .s1,
  .codehilite .s2,
  .codehilite .s  {{ color: #98c379 }}
  .codehilite .mi {{ color: #d19a66 }}
  .codehilite .bp {{ color: #56b6c2 }}
  .codehilite .o  {{ color: #56b6c2 }}
  .codehilite .p  {{ color: #abb2bf }}
  .codehilite .n  {{ color: #e06c75 }}

  /* ── Footer ── */
  .footer {{
    margin-top: 40px;
    border-top: 1px solid #e5e7eb;
    padding-top: 12px;
    text-align: center;
    font-size: 8pt;
    color: #9ca3af;
  }}
</style>
</head>
<body>

<div class="cover">
  <div class="brand">Nova Testing Platform</div>
  <h1>AI Code Analysis Report</h1>
  <div class="sub">Generated on {ts}</div>
</div>

<div class="content">
  {html_content}
  <div class="footer">
    Generated by <strong>Nova</strong> &mdash; AI-Powered Automated Testing Platform
  </div>
</div>

</body>
</html>"""

    options = {
        'page-size':      'A4',
        'margin-top':     '0in',
        'margin-right':   '0in',
        'margin-bottom':  '0.5in',
        'margin-left':    '0in',
        'encoding':       'UTF-8',
        'print-media-type': '',
        'no-outline':     None,
    }

    try:
        pdf = pdfkit.from_string(full_html, False, options=options)
        return pdf
    except Exception as e:
        print(f"[pdfkit] Error generating code analysis PDF: {e}")
        return None