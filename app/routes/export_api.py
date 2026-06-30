"""
Export API routes — PDF and Excel report download endpoints.
"""
import json
import tempfile

from flask import Blueprint, request, jsonify, send_file, current_app, redirect, url_for, flash
from flask_login import login_required, current_user
from fpdf import FPDF
from openpyxl import Workbook

from app.models import TestRun
from app.services.report_service import (
    generate_pdf_report,
    generate_inline_pdf,
    generate_excel_report,
    generate_code_analysis_pdf,
)
from app.utils import TestVersePDF, sanitize_text

export_bp = Blueprint('export', __name__)


@export_bp.route('/export/pdf', methods=['POST'])
@login_required
def export_pdf():
    data    = request.get_json() or {}
    summary = data.get("summary", "Test Execution Report")
    results = data.get("results", [])
    url     = data.get("url")
    tmp     = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp.close()
    try:
        generate_inline_pdf(summary, results, tmp.name, url=url)
    except Exception as e:
        current_app.logger.exception("Inline PDF error")
        return jsonify({"error": str(e)}), 500
    return send_file(tmp.name, as_attachment=True, download_name="Test_Report.pdf")


@export_bp.route('/export/excel', methods=['POST'])
@login_required
def export_excel():
    results = (request.get_json() or {}).get("results", [])
    wb = Workbook()
    ws = wb.create_sheet("Test Results", 0)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    ws.append(["Step", "Command", "Status", "Details"])
    for row in results:
        ws.append([row.get("step", ""), row.get("command", ""),
                   row.get("status", ""), row.get("details", "")])
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name="Test_Report.xlsx")


@export_bp.route('/download-history/<int:test_run_id>')
@login_required
def download_history(test_run_id):
    tr = TestRun.query.get_or_404(test_run_id)
    if tr.user_id != current_user.id:
        flash('Unauthorized.', 'danger')
        return redirect(url_for('main.history'))
    results  = json.loads(tr.results_json)
    filepath = tempfile.mktemp(suffix=".pdf")
    
    # Generate dynamic execution summary
    failed_steps = [r for r in results
                    if str(r.get("status", "")).lower() in ("failed", "visual mismatch", "visuals mismatch", "visual_mismatch")]
    total_steps  = len(results)
    summary      = (f"Executed {total_steps} step(s) on {tr.url_tested}. " +
                    (f"{len(failed_steps)} step(s) failed." if failed_steps else "All steps succeeded."))
    
    try:
        generate_pdf_report(tr.url_tested, results, filepath, summary=summary)
    except Exception:
        _simple_pdf(tr, results, filepath)
    return send_file(filepath, as_attachment=True,
                     download_name=f'Test_Report_{test_run_id}.pdf')


@export_bp.route('/download-excel/<int:test_run_id>')
@login_required
def download_excel_history(test_run_id):
    tr = TestRun.query.get_or_404(test_run_id)
    if tr.user_id != current_user.id:
        flash('Unauthorized.', 'danger')
        return redirect(url_for('main.history'))
    results = json.loads(tr.results_json)
    try:
        data = generate_excel_report(results, tr.url_tested, current_user.name, test_run_id)
        tmp  = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(data); tmp.close()
    except Exception:
        tmp  = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb   = Workbook()
        ws   = wb.active
        ws.append(["Step", "Command", "Status", "Details"])
        for r in results:
            ws.append([r.get("step"), r.get("command"), r.get("status"), r.get("details")])
        wb.save(tmp.name); tmp.close()
    return send_file(tmp.name, as_attachment=True,
                     download_name=f'Test_Report_{test_run_id}.xlsx')


@export_bp.route('/download_code_analysis_pdf', methods=['POST'])
@login_required
def download_code_analysis_pdf():
    data = request.get_json() or {}
    md_text = data.get('ai_result_markdown', '')
    code_text = data.get('code', '').strip()
    
    if not md_text:
        return jsonify({"error": "No analysis data"}), 400
        
    if code_text:
        md_text = f"# Code Analyzed\n```python\n{code_text}\n```\n\n{md_text}"
    try:
        pdf_bytes = generate_code_analysis_pdf(md_text)
        if pdf_bytes:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            tmp.write(pdf_bytes); tmp.close()
            return send_file(tmp.name, as_attachment=True,
                             download_name='AI_Code_Analysis_Report.pdf')
        # pdfkit not available — fallback
        raise RuntimeError("pdfkit unavailable")
    except Exception as e:
        current_app.logger.warning(f"Code analysis PDF fallback: {e}")
        pdf = TestVersePDF(report_title="AI Code Analysis Report")
        pdf.add_page()
        
        # Styled title
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(79, 70, 229)  # Iris brand primary
        pdf.cell(pdf.epw, 10, "AI Code Analysis Report", new_x="LMARGIN", new_y="NEXT", align="L")
        pdf.ln(5)
        
        in_code_block = False
        
        md_text = md_text.replace('\r\n', '\n').replace('\r', '\n')
        for line in md_text.split('\n'):
            safe = sanitize_text(line)
            stripped = safe.strip()
            
            if stripped.startswith("```"):
                in_code_block = not in_code_block
                pdf.ln(2)
                continue
                
            if in_code_block:
                pdf.set_font("Courier", "", 9.5)
                pdf.set_text_color(31, 41, 55)
                pdf.set_fill_color(243, 244, 246)
                pdf.multi_cell(pdf.epw, 5, safe, fill=True, new_x="LMARGIN", new_y="NEXT")
                continue
                
            # Render standard markdown element outside code block
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(17, 24, 39)
            
            if not stripped:
                pdf.ln(2)
                continue
                
            if stripped.startswith("#"):
                num_hashes = len(stripped) - len(stripped.lstrip('#'))
                clean_title = stripped.lstrip('#').strip()
                if num_hashes == 1:
                    pdf.ln(4)
                    pdf.set_font("Helvetica", "B", 14)
                    pdf.set_text_color(79, 70, 229)
                    pdf.multi_cell(pdf.epw, 7, clean_title, new_x="LMARGIN", new_y="NEXT")
                elif num_hashes == 2:
                    pdf.ln(3)
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.set_text_color(17, 24, 39)
                    pdf.multi_cell(pdf.epw, 6, clean_title, new_x="LMARGIN", new_y="NEXT")
                else:
                    pdf.ln(2)
                    pdf.set_font("Helvetica", "B", 11)
                    pdf.set_text_color(107, 114, 128)
                    pdf.multi_cell(pdf.epw, 6, clean_title, new_x="LMARGIN", new_y="NEXT")
            elif stripped.startswith("- ") or stripped.startswith("* "):
                pdf.set_x(15)
                pdf.cell(4, 5.5, chr(149))
                pdf.multi_cell(pdf.epw - 5, 5.5, stripped[2:], new_x="LMARGIN", new_y="NEXT")
            else:
                pdf.multi_cell(pdf.epw, 5.5, safe, new_x="LMARGIN", new_y="NEXT")
                
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        pdf.output(tmp.name)
        return send_file(tmp.name, as_attachment=True,
                         download_name='AI_Code_Analysis_Report.pdf')


def _simple_pdf(tr, results: list, filepath: str):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Test Report", ln=True, align="C")
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"URL: {tr.url_tested}", ln=True)
    pdf.cell(0, 8, f"Status: {tr.status}", ln=True)
    pdf.cell(0, 8, f"Date: {tr.timestamp.strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(4)
    for r in results:
        pdf.cell(0, 6, f"Step {r.get('step')}: [{r.get('status')}] {r.get('command', '')}", ln=True)
    pdf.output(filepath)
